import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.90
    corrected_cell = sheet.cell(row, 4)
    corrected_cell.value = corrected_price

chart = BarChart()
values = Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=sheet.max_row + 1)
chart.add_data(values)\
sheet.add_chart(chart, "E2")

wb.save('transactions2.xlsx')




